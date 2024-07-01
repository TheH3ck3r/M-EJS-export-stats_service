from ..models.stats import Stats, absence_type_view, months_in_semester, Item, discipline_type_view, CommonStats, MonthStats
import uuid
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import (Alignment)
from openpyxl.utils import get_column_interval, get_column_letter


async def export_stats(stats: Stats, common_stats: list[CommonStats], month_stats: list[MonthStats]):
    column_width = len(max(stats.students.values(), key=len)) * 1.2
    wb = Workbook()
    wb.active.title = "Основная информация"
    common_stats_table = _create_common_stats_table(
        common_stats, stats.disciplines, stats.students, month_stats)
    for row in common_stats_table:
        wb.active.append(row)
    for row in wb.active.columns:
        for cell in row:
            alignment = Alignment(
                horizontal="center", vertical="center")
            cell.alignment = alignment
    for cell in wb.active[1][1:-5]:
        alignment = Alignment(
            text_rotation=90, horizontal="center", vertical="center")
        cell.alignment = alignment
    wb.active.column_dimensions["A"].width = column_width

    for col in get_column_interval("B", get_column_letter((len(stats.disciplines) + 1))):
        wb.active.column_dimensions[col].width = 4

    for discipline_id, discipline_name in stats.disciplines.items():
        ws = wb.create_sheet(discipline_name[0])
        table = _create_stats_table(stats, discipline_id)
        for row in table:
            ws.append(row)
        for row in ws.columns:
            for cell in row:
                alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.alignment = alignment
        ws.column_dimensions["A"].width = column_width
    file_name = f"{uuid.uuid4()}.xlsx"
    wb.save(f"dist/{file_name}")
    return file_name


def _create_common_stats_table(stats: list[CommonStats], disciplines: dict[uuid.UUID, str], students: dict[uuid.UUID, str], month_stats: list[MonthStats]):
    table_head = ["ФИО Студента"]
    for discipline_id, discipline_name in disciplines.items():
        table_head.append(discipline_name[1])
    for month in months_in_semester(1):
        table_head.append(month)
    table_head.append("Всего")

    table = [table_head]
    for student_id, student_fullname in students.items():
        row = [student_fullname]
        for discipline_id, discipline_name in disciplines.items():
            data = [j for j in stats if j.student_id ==
                    student_id and j.discipline_id == discipline_id]
            if data:
                row.append(data[0].count)
            else:
                row.append("0")

        for month in months_in_semester(1):
            row.append(_common_absence_count_at_month(
                stats=month_stats, student_id=student_id, month=month))
        row.append(_common_absence_count_at_month(
            stats=month_stats, student_id=student_id))

        table.append(row)
    return table


def _common_absence_count_at_month(stats: list[MonthStats], student_id: uuid.UUID, month: str | None = None):
    if month == None:
        absence_count = 0
        for data in stats:
            if data.student_id == student_id:
                absence_count += data.count
        return absence_count
    else:
        month_number = _month_to_int(month) - 1
        for data in stats:
            if data.student_id == student_id and data.date.month == month_number:
                return data.count
        return 0


def _month_to_int(month: str):
    return {
        'Январь': 1,
        'Февраль': 2,
        'Март': 3,
        'Апрель': 4,
        'Май': 5,
        'Июнь': 6,
        'Июль': 7,
        'Август': 8,
        'Сентябрь': 9,
        'Октябрь': 10,
        'Ноябрь': 11,
        'Декабрь': 12
    }[month]


def _absence_count_at_month(jobs: list[Item], student_id: uuid.UUID, month: str | None = None):
    absence_count = 0
    if month == None:
        for job in jobs:
            if job.student_id == student_id and job.absence_type == 0:
                absence_count += 1
    else:
        month_number = _month_to_int(month)
        for job in jobs:
            if job.student_id == student_id and job.absence_type == 0 and job.date.month == month_number:
                absence_count += 1
    return absence_count


def _create_stats_table(stats: Stats, discipline_id: uuid.UUID):
    table_head = ["ФИО Студента"]
    unique_jobs = {}
    jobs = list(filter(lambda j: j.discipline_id ==
                discipline_id, stats.items))
    for job in jobs:
        if job.discipline_id == discipline_id:
            key = str(job.date) + job.start_at
            if key not in unique_jobs:
                unique_jobs[key] = job
    unique_jobs = list(unique_jobs.values())
    for job in sorted(unique_jobs, key=lambda j: j.date):
        table_head.append(
            f"{job.date.strftime('%d.%m')}\n{discipline_type_view(job.discipline_type)}")
    for month in months_in_semester(1):
        table_head.append(month)
    table_head.append("Всего")

    table = [table_head]
    for student_id, student_fullname in stats.students.items():
        row = [student_fullname]
        for unique_job in sorted(unique_jobs, key=lambda j: j.date):
            job = [j for j in jobs if unique_job.date ==
                   j.date and student_id == j.student_id]
            if job:
                row.append(absence_type_view(job[0].absence_type))
            else:
                row.append(" ")
        for month in months_in_semester(1):
            row.append(_absence_count_at_month(
                jobs=jobs, student_id=student_id, month=month))
        row.append(_absence_count_at_month(
            jobs=jobs, student_id=student_id))
        table.append(row)
    return table
